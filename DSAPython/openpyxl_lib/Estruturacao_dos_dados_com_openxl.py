from openpyxl import Workbook, load_workbook
import os
import glob
import pandas as pd
import csv

counter = 0

#csv_files_2 = glob.glob(os.path.join('C:\\Users\\adria\\OneDrive\\Documentos\\DSAPython\\openpyxl_lib\\files_csv','*.csv'))
#glob.glob() realiza uma varredura nos diretórios listando os arquivos que (com estensões especificadas ou não - '\\files_csv\\*.csv').
#Para incluir um grupo de diretórios de um determinado endereço pode-se complementar desta forma - '\\files_csv\\**\\*.csv'
csv_files_2 = glob.glob('C:\\Users\\adria\\OneDrive\\Documentos\\DSAPython\\openpyxl_lib\\files_csv\\*.csv', recursive=True)
#Verificando o número de arquivos dentro do diretório em que os arquivos de registro de temperaturas por cidade 
qtd_arq = os.listdir('C:\\Users\\adria\\OneDrive\\Documentos\\DSAPython\\openpyxl_lib\\files_csv')
x = len(qtd_arq)

for file in csv_files_2:
    #split() separa a string segundo os parâmetros definidos string.split(separator, maxsplit)
    #Desta forma é possível modificar o path para o novo diretório com a nova extensão, conservando o nome do arquivo.
    name_1 = file.split('files_csv')
    name_2 = name_1[1].split('-2021')
    new_name_file = name_1[0] + "files_xls" + name_2[0] + "-2021.xlsx"
    print(new_name_file)
    #Realiza a cópia de todo o conteúdo da planilha e adiciona em uma nova planilha e em seguida salva as informações com o
    #novo nome e caminho definidos acima.
    wb = Workbook()
    ws = wb.active
    with open (file) as f :
        reader = csv.reader(f, delimiter = ';')
        for row in reader:
            ws.append(row)
    wb.save(new_name_file)
    x = x-1
    if x == 0:
        break
    
#Todo esse processo de conversão dos arquivos com extensão .csv para .xlsx se deve ao fato do pacote "openpyxl" não suportar
#a extensão .csv. Se houver a tentativa de manipular os arquivos com essa extensão, será gerado a mensagem de erro abaixo:
#raise InvalidFileException(msg)
#openpyxl.utils.exceptions.InvalidFileException: openpyxl does not support .csv file format, please check you can open it with
#Excel first. Supported formats are: .xlsx,.xlsm,.xltx,.xltm


#-------------------------------------------------------
#Voltando a função glob.glob() para acessar os novos arquivos com a extensão .xlsx criado anteriormente
csv_files = glob.glob('C:\\Users\\adria\\OneDrive\\Documentos\\DSAPython\\openpyxl_lib\\files_xls\\*.xlsx', recursive=True)
qtd_arq2 = os.listdir('C:\\Users\\adria\\OneDrive\\Documentos\\DSAPython\\openpyxl_lib\\files_xls')
y = len(qtd_arq2)
for f in csv_files :
    #Lendo o arquivo acessado pela função glob.glob()
    wb = load_workbook(f)
    #Fazendo uma varredura em todas as abas existente do arquivo.
    for sheet in wb:
        aba = sheet
        ws = wb.active
        #Neste passo, é possível organizar os dados da planilha para um formato mais adequado de uma tabela de banco de dados
        #copiando valores das células existente para as respectivas colunas (consideradas agora como nomes de colunas).
        ws['T9'] = ws['A1'].value
        ws['U9'] = ws['A2'].value
        ws['V9'] = ws['A3'].value
        ws['W9'] = ws['A4'].value
        ws['X9'] = ws['A5'].value
        ws['Y9'] = ws['A6'].value
        ws['Z9'] = ws['A7'].value
        ws['AA9'] = ws['A8'].value
        #Verifica-se a quantidade de células da coluna "A" que contenham quaisquer valores. Essa quantidade será utilizada
        #para limitar o cíclo da estrutura de repetição somente ao necessário 
        sizes = len(ws['a'])
        #Essa variável indica que é preciso considerar os valores a partir da linha 10.
        lineT = 10
        #A estrutura de repeticação abaixo garante que os valores nas células especificadas sejam repetidas nas colunas
        #que foram anteriormente pré-definidas para receberem esses valores

        for row in ws['A10:A9999'] :
            #Avalues = ws.cell(lineT,1).value
                   #linha,coluna        #Célula
            ws.cell(lineT,20).value = ws['B1'].value
            ws.cell(lineT,21).value = ws['B2'].value
            ws.cell(lineT,22).value = ws['B3'].value

            ws.cell(lineT,23).value = ws['B4'].value
            ws.cell(lineT,24).value = ws['B5'].value
            ws.cell(lineT,25).value = ws['B6'].value
            ws.cell(lineT,26).value = ws['B7'].value
            ws.cell(lineT,27).value = ws['B8'].value
            lineT = lineT + 1
            sizes = sizes - 1
            if sizes == 0:
                break
        #Exclui as 8 primeiras linhas
                    #index(número da linha),#quantidade de linhas a serem deletadas
        ws.delete_rows(1,8)
    print(f)
    print(str(y) + ' de ' + str(len(qtd_arq2)))
    y = y-1
    wb.save(f)
    if y == 0:
        print('FINISHED!!!!!!!!!!!!!!!!!!!')
        break

#--------------------------------------------------------------------------------------
#Considerando que trata-se de aproximadamente 800 arquivos contendo entre 8000 e 9000 linhas, cada.
#Como opção foi adotado a 























            

